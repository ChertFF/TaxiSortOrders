from datetime import date

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

#1. Подготовка.... видоизменяем и удаляем файл с прошлой выгрузки
#Удаляем данные с прошлой выгрузки
try:
    os.remove('orders-out.xlsx')
except:
    pass

#Открываем книгу и добавляем первой строкой заголовки для таблицы по заказам
wb = openpyxl.load_workbook('orders-in.xlsx')
sheet = wb.active

#Устанавлием названия заголовков
header_values = ['Время', 'Населенный пункт', 'Адрес', 'ФИО', 'Номер телефона', 'Номер заказа', 'Номер машины']

#Удаляем те строки, которые содержать те же значения, что в верхнем массиве
#Нужно, чтобы удалить дубликаты, которые могли быть при копировании
rows_to_delete = []
for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), start=1):
    if any(value in row for value in header_values):
        rows_to_delete.append(row_number)
for row in reversed(rows_to_delete):
    sheet.delete_rows(row)

#Добавляем заголовки таблицы первой строкой
sheet.insert_rows(1)
for col_num, value in enumerate(header_values, 1):
    col_letter = get_column_letter(col_num)
    sheet[f'{col_letter}1'] = value

#Сохраняем изменения в исходном файле
wb.save('orders-in.xlsx')

#Закрываем книгу
wb.close()

#2. Используем pandas для группировки по машинам
# Чтение данных из Excel файла
df = pd.read_excel('orders-in.xlsx')

#Находим количество изначальных строк в таблице (-1, ибо отнимаем строку с заголовком таблицы)
count_original = df.shape[0]

# Замена символов в номере заказа
df['Номер заказа'] = pd.to_numeric(df['Номер заказа'], errors='coerce')

# Сортировка данных по номеру машины
df = df.sort_values('Номер заказа')

# Группируем заказы по номеру машины
grouped_df = df.groupby('Номер заказа')

# Создаем новый DataFrame для объединенных заказов
merged_df = pd.DataFrame(columns=df.columns)

# Счетчик строк с номером заказа сколько вышло по факту
count_default = 0

# Объединяем заказы по машинам
for _, group in grouped_df:
    merged_df = pd.concat([merged_df, pd.DataFrame([{'Время': f"Машина на {group['Время'].iloc[0]}", 'Населенный пункт': '', 'Адрес': '', 'ФИО': '', 'Номер телефона': '', 'Номер заказа': '', 'Номер машины': ''}]), group], ignore_index=True)
    count_default += len(group)

# Выводим таблицу
print(merged_df)

#Сверяем значения выгрузки до и после
if (count_default == count_original):
    print(f'Количество изначальных строк заказов совпадает с количестом выгруженных: {count_default}/{count_original}\n'
          f'Хорошего тебе дня, счастья, здоровья, денег побольше тебе и твоим близким ♥')
else: print(f"АЛАРМ! ЧТО-ТО ПОШЛО НЕ ТАК!\n"
            f"Выгружено меньше заказов, чем было изначально:{count_default}/{count_original}\n"
            f"Перепроверь данные, если есть заказы с дробью, то замени точку в них на запятую перед вставкой в соседний файлик")

# Сохраняем данные в файл 'orders-out.xlsx'
merged_df.to_excel('orders-out.xlsx', index=False)


#3. Видоизменяем готовый файл - наводим красоту
# Открываем книгу
wb = openpyxl.load_workbook('orders-out.xlsx')

# Выбираем активный лист
ws = wb.active

# # Скрываем столбец F
ws.column_dimensions['F'].hidden = True

#3.1 Объединяем ячейки начинающиеся на "Машина на..."
# Создаем список для хранения диапазонов ячеек, которые нужно объединить
ranges_to_merge = []

# Обходим строки с первого до шестого столбца
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
    if any("Машина на" in str(cell.value) for cell in row):
        # Получаем диапазон ячеек для объединения
        start_row = row[0].row
        end_row = row[0].row
        start_column = 1
        end_column = 7
        range_to_merge = f"A{start_row}:G{end_row}"
        ranges_to_merge.append(range_to_merge)

        # Изменяем высоту объединенных ячеек на 30 пикселей
        ws.row_dimensions[start_row].height = 25

        # Выделяем текст жирным
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="C2E0C6", end_color="C2E0C6", fill_type="solid")

# Объединяем ячейки начинающиеся на "Машина на..."
for range_to_merge in ranges_to_merge:
    ws.merge_cells(range_to_merge)


# Установка шрифта и размера шрифта для всей книги
font = Font(name='Calibri', size=12)
ws.font = font

# Установка выравнивания по центру для столбцов A, B, C, D, E
align_center = Alignment(horizontal='center', vertical='center')
for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    for row in range(1, ws.max_row + 1):
        cell = ws[column + str(row)]
        cell.alignment = align_center

# Выравниваем все столбцы по ширине
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.1
    ws.column_dimensions[column[0].column_letter].width = adjusted_width

# Установка границ для всех ячеек, начиная с первой строки
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))
for row in range(1, ws.max_row + 1):
    for column in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=column)
        cell.border = border

# Заменяем первую строку на объединенную ячейку с надписью "Итоговая доставка на [Сегодняшняя дата]"
first_row = ws[1:1]
first_row[0].value = f"Итоговая доставка на {date.today()}"

#Объединяем строки
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
#Форматируем так, как ячейки с текстом "Машина на...": полужирны, заливка и высота ячейки
first_row[0].font = Font(bold=True)
first_row[0].fill = PatternFill(start_color="C2E0C6", end_color="C2E0C6", fill_type="solid")
ws.row_dimensions[1].height = 25

# Сохраняем изменения
wb.save('orders-out.xlsx')

# Закрываем книгу
wb.close()

# Добавьте эту строку в конце кода
input(f"Нажмите Enter для выхода...")